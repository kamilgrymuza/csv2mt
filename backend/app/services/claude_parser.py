"""
Claude AI Document Parser Service

This service uses Claude AI to automatically extract transaction data from
various document formats (CSV, PDF, XLS/XLSX) without needing bank-specific parsers.

The AI identifies transaction data, extracts it into a standardized format,
and converts it to MT940 format.
"""

import base64
import csv
import io
import json
import re
from datetime import datetime
from typing import List, Dict, Any, Optional, BinaryIO
from anthropic import Anthropic
import PyPDF2
import openpyxl

from ..config import settings
from .encoding_detector import decode_file_content, EncodingDetectionError


class EmptyStatementError(ValueError):
    """
    Exception raised when a statement file contains no valid transaction data.

    This is an expected error condition (not a bug) and should NOT be logged to Sentry.
    It indicates that the user uploaded a file that doesn't contain parseable transaction data.
    """
    pass

# Import OpenAI if available
try:
    from openai import OpenAI
    OPENAI_AVAILABLE = True
except ImportError:
    OPENAI_AVAILABLE = False


class Transaction:
    """Standardized transaction representation"""

    def __init__(
        self,
        date: str,
        amount: float,
        description: str,
        transaction_type: str = "UNKNOWN",
        reference: Optional[str] = None,
        balance: Optional[float] = None
    ):
        self.date = date
        self.amount = amount
        self.description = description
        self.transaction_type = transaction_type
        self.reference = reference
        self.balance = balance

    def to_dict(self) -> Dict[str, Any]:
        return {
            "date": self.date,
            "amount": self.amount,
            "description": self.description,
            "transaction_type": self.transaction_type,
            "reference": self.reference,
            "balance": self.balance
        }


class ClaudeDocumentParser:
    """Service for parsing documents using AI (Claude or OpenAI)"""

    def __init__(self, model: Optional[str] = None, disable_cache: bool = False):
        """Initialize parser with specified AI model

        Args:
            model: AI model to use (claude-sonnet, claude-haiku, gpt-4o)
                   If None, uses settings.ai_model
            disable_cache: If True, adds unique identifier to break AI response caching
                          Useful for testing to ensure fresh responses
        """
        self.model = model or settings.ai_model
        self.disable_cache = disable_cache

        if self.model.startswith("claude"):
            if not settings.anthropic_api_key:
                raise ValueError("ANTHROPIC_API_KEY not configured")
            self.client = Anthropic(api_key=settings.anthropic_api_key)
            self.openai_client = None
        elif self.model.startswith("gpt"):
            if not OPENAI_AVAILABLE:
                raise ValueError("OpenAI package not installed. Run: pip install openai")
            if not settings.openai_api_key:
                raise ValueError("OPENAI_API_KEY not configured")
            self.openai_client = OpenAI(api_key=settings.openai_api_key)
            self.client = None
        else:
            raise ValueError(f"Unsupported model: {self.model}")

    def parse_document(
        self,
        file_content: BinaryIO,
        filename: str,
        account_number: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Parse a document and extract transaction data using Claude AI

        Args:
            file_content: Binary file content
            filename: Original filename (used to determine file type)
            account_number: Optional account number for MT940 generation

        Returns:
            Dict containing transactions and metadata
        """
        file_type = self._get_file_type(filename)

        # Extract text content based on file type
        if file_type == "csv":
            text_content = self._extract_csv_content(file_content)
            result = self._parse_with_claude_text(text_content, account_number, file_type="csv")
        elif file_type == "pdf":
            # Send PDF directly to Claude (preserves visual layout)
            result = self._parse_pdf_with_vision(file_content, account_number)
        elif file_type in ["xls", "xlsx"]:
            text_content = self._extract_excel_content(file_content)
            # Pass file_content so we can retry with CSV conversion if needed
            result = self._parse_with_claude_text(text_content, account_number, file_type="excel", file_content=file_content)
        else:
            raise ValueError(f"Unsupported file type: {file_type}")

        # Store original filename in metadata for Field 20 reference generation
        if "metadata" not in result:
            result["metadata"] = {}
        result["metadata"]["source_filename"] = filename

        return result

    def _get_file_type(self, filename: str) -> str:
        """Determine file type from filename"""
        extension = filename.lower().split('.')[-1]
        if extension in ['xls', 'xlsx']:
            return 'xlsx'
        return extension

    def _extract_csv_content(self, file_content: BinaryIO) -> str:
        """
        Extract text from CSV file with automatic encoding detection.

        Uses charset-normalizer to reliably detect the file encoding.
        Fails explicitly if encoding cannot be detected with sufficient confidence
        rather than guessing, which ensures accurate data extraction.

        For CSV files, we use a lower confidence threshold (50%) because:
        - CSV files are often small with limited text
        - Limited text makes statistical detection less reliable
        - Even with lower confidence, charset-normalizer is still more accurate
          than blind fallback attempts

        Raises:
            EncodingDetectionError: If encoding cannot be detected reliably
        """
        content = file_content.read()

        # Use centralized encoding detection with 50% minimum confidence for CSV
        # (CSV files are typically small, making statistical detection harder)
        #
        # Prioritize Eastern European encodings (windows-1250, iso-8859-2) over
        # Western European ones (cp1252, latin-1) when confidence scores are close,
        # as many banking systems in Central/Eastern Europe use these encodings
        text, detected_encoding = decode_file_content(
            content,
            min_confidence=0.5,
            filename="CSV file",
            prioritize_encodings=['windows-1250', 'cp1250', 'iso-8859-2']
        )

        print(f"✓ Detected CSV encoding: {detected_encoding}")
        return text

    def _parse_pdf_with_vision(self, file_content: BinaryIO, account_number: Optional[str] = None) -> Dict[str, Any]:
        """
        Parse PDF by sending it directly to Claude with vision.
        This preserves the visual layout and table structure.

        Args:
            file_content: Binary PDF content
            account_number: Optional account number

        Returns:
            Dict with transactions and metadata
        """
        print("📄 Processing PDF with Claude vision API...")

        # Read PDF as bytes
        pdf_bytes = file_content.read()
        pdf_base64 = base64.b64encode(pdf_bytes).decode('utf-8')

        # Get page count for analytics
        import io
        pdf_reader = PyPDF2.PdfReader(io.BytesIO(pdf_bytes))
        file_page_count = len(pdf_reader.pages)
        print(f"   PDF has {file_page_count} page(s)")

        prompt = """You are a financial document parser. Analyze this bank statement PDF and extract ALL transaction data.

Look at the visual table structure carefully and extract every transaction.

Output format - CSV with properly quoted fields:

METADATA,account_number,currency,statement_start_date,statement_end_date,opening_balance,closing_balance
TXN,date,amount,"description",transaction_type,reference,balance
TXN,date,amount,"description",transaction_type,reference,balance
...

Example output:
METADATA,PL12345678,PLN,2024-01-01,2024-01-31,1000.00,500.00
TXN,2024-01-15,-100.50,"Amazon Purchase",DEBIT,REF123,899.50
TXN,2024-01-16,500.00,"Salary Deposit",CREDIT,SAL2024,1399.50

Field details:
- account_number: remove ALL whitespace from IBANs
- currency: 3-letter ISO code (USD/EUR/GBP/PLN) - look for symbols €,$,£,zł
- date: YYYY-MM-DD format
- amount: negative for debits/money out, positive for credits/money in
- description: MUST be quoted, combine relevant columns (see CRITICAL rules below)
- transaction_type: CREDIT or DEBIT
- reference: transaction reference number (empty if not available)
- balance: account balance after transaction (empty if not available)

CRITICAL - Description Field Extraction Rules (EXACT CHARACTER-BY-CHARACTER COPYING):
1. Extract the COMPLETE, FULL text EXACTLY as it appears - character-by-character
2. Include ALL text visible in the cells - do NOT summarize, abbreviate, or omit any parts
3. PRESERVE EXACT FORMATTING including:
   - Leading and trailing spaces (but normalize to single space after combining)
   - Slashes, hyphens, and all punctuation EXACTLY as shown in PDF
   - Special characters like "/", "-", ",", ".", etc. in their EXACT positions
4. If a table has separate columns for:
   - Counterparty/payee (who): Extract the ENTIRE cell content EXACTLY
   - Title/description (what): Extract the ENTIRE cell content EXACTLY
   - Combine them as: "counterparty - title" (with space-dash-space separator)
5. ONLY remove 26-digit bank account numbers (IBAN-like numbers) from the description
6. KEEP all other information EXACTLY as shown:
   - NIP/tax ID numbers - preserve EXACTLY: "/NIP/" or "NIP/" as shown
   - Invoice numbers - preserve EXACTLY
   - Reference codes - preserve EXACTLY with all slashes: "/TI/", "/OKR/", etc.
   - ALL other details visible in the cells
7. PAY SPECIAL ATTENTION to slashes before codes like NIP, TI, OKR:
   - If PDF shows "/NIP/123", write "/NIP/123" (with leading slash)
   - If PDF shows "NIP/123", write "NIP/123" (without leading slash)
   - Copy EXACTLY what you see - do not add or remove slashes
8. Be DETERMINISTIC - extract EXACTLY the same characters every time you process the same PDF

Examples of EXACT extraction (note: copy every character including slashes exactly as shown):
- If PDF shows: "... - /NIP/5862147548/..." → write: "... - /NIP/5862147548/..." ✓
- If PDF shows: "... - NIP/5862147548/..." → write: "... - NIP/5862147548/..." ✓
- If PDF shows: "... - /TI/N123/OKR/0/..." → write: "... - /TI/N123/OKR/0/..." ✓
- "PROWIZJE I OPŁATY-POZOSTAŁE I/616 - opł. za opinię bankową" ✓
- "Adam Cioczek - 1 rata wynagrodzenia za przeniesienie praw do treatmentu filmu Przejście z godnie z rachunkiem z dnia 12.09.2025" ✓

IMPORTANT:
- Return ONLY the CSV format, no other text or explanations
- ALWAYS quote description fields
- Extract ALL transactions from ALL pages
- Use the visual table structure to correctly identify columns
- Be DETERMINISTIC - always extract the same information from the same cells"""

        # Call Claude with PDF document
        if self.client:  # Claude (Anthropic)
            model_name = "claude-sonnet-4-5-20250929" if self.model == "claude-sonnet" else "claude-3-5-haiku-20241022"
            response = self.client.messages.create(
                model=model_name,
                max_tokens=8192,
                temperature=0.0,  # Use deterministic mode for consistent results
                messages=[{
                    "role": "user",
                    "content": [
                        {
                            "type": "document",
                            "source": {
                                "type": "base64",
                                "media_type": "application/pdf",
                                "data": pdf_base64
                            }
                        },
                        {
                            "type": "text",
                            "text": prompt
                        }
                    ]
                }]
            )
            response_text = response.content[0].text
            input_tokens = response.usage.input_tokens
            output_tokens = response.usage.output_tokens
        else:
            raise ValueError("PDF vision parsing currently only supported with Claude (Anthropic API)")

        # Parse CSV response
        try:
            result = self._parse_pipe_delimited_response(response_text)

            # Add token usage and parsing method to metadata
            if "metadata" not in result:
                result["metadata"] = {}
            result["metadata"]["input_tokens"] = input_tokens
            result["metadata"]["output_tokens"] = output_tokens
            result["metadata"]["parsing_method"] = "pdf_vision"
            result["metadata"]["file_page_count"] = file_page_count

            print(f"   ✓ Extracted {len(result.get('transactions', []))} transactions")
            print(f"   ✓ Tokens: {input_tokens:,} input, {output_tokens:,} output")

            return result

        except Exception as e:
            print(f"   ✗ Failed to parse CSV response: {e}")
            print(f"   Response: {response_text[:500]}")
            raise ValueError(f"AI returned invalid CSV: {str(e)}")

    def _extract_pdf_content(self, file_content: BinaryIO) -> str:
        """
        Extract text from PDF file (fallback method, not used when vision API is available)
        """
        pdf_reader = PyPDF2.PdfReader(file_content)
        text_parts = []
        for page in pdf_reader.pages:
            text_parts.append(page.extract_text())
        return "\n".join(text_parts)

    def _extract_excel_content(self, file_content: BinaryIO) -> str:
        """Extract text from Excel file (tab-delimited format)"""
        workbook = openpyxl.load_workbook(file_content, data_only=True)
        text_parts = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text_parts.append(f"=== Sheet: {sheet_name} ===")

            for row in sheet.iter_rows(values_only=True):
                # Convert all values to strings and join
                row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                if row_text.strip():
                    text_parts.append(row_text)

        return "\n".join(text_parts)

    def _convert_excel_to_csv(self, file_content: BinaryIO) -> str:
        """
        Convert Excel file to proper CSV format (comma-delimited).

        This is used as a retry mechanism when tab-delimited format fails.
        CSV format is more standard and works better with the format spec approach.

        Args:
            file_content: Excel file binary content

        Returns:
            CSV-formatted string
        """
        workbook = openpyxl.load_workbook(file_content, data_only=True)
        csv_parts = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            csv_parts.append(f"=== Sheet: {sheet_name} ===")

            # Use CSV writer for proper escaping and formatting
            output = io.StringIO()
            writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL)

            for row in sheet.iter_rows(values_only=True):
                # Convert None to empty string, everything else to string
                row_values = [str(cell) if cell is not None else "" for cell in row]
                # Skip completely empty rows
                if any(val.strip() for val in row_values):
                    writer.writerow(row_values)

            csv_parts.append(output.getvalue().strip())

        return "\n".join(csv_parts)

    def _chunk_text_by_lines(self, text_content: str, lines_per_chunk: int = 100) -> List[str]:
        """
        Split text content into chunks by lines (for large files)

        Args:
            text_content: The full text content
            lines_per_chunk: Number of lines per chunk (targets ~80-90 transactions = ~6000 output tokens with 8192 limit)

        Returns:
            List of text chunks
        """
        lines = text_content.split('\n')

        # Keep header lines (first 20 lines usually contain metadata)
        header_lines = lines[:20]
        data_lines = lines[20:]

        # If file is small enough, return as single chunk
        if len(data_lines) <= lines_per_chunk:
            return [text_content]

        # Split into chunks
        chunks = []
        for i in range(0, len(data_lines), lines_per_chunk):
            chunk_data = data_lines[i:i + lines_per_chunk]
            # Each chunk includes header + data portion
            chunk_text = '\n'.join(header_lines + chunk_data)
            chunks.append(chunk_text)

        return chunks

    def _parse_with_claude_text(
        self,
        text_content: str,
        account_number: Optional[str] = None,
        file_type: str = "csv",
        file_content: Optional[BinaryIO] = None
    ) -> Dict[str, Any]:
        """
        Parse text content using appropriate method based on file type.

        For CSV/Excel: Two-step approach (format spec + Python parsing)
        For PDF: Direct AI extraction (PDFs have multi-line formats)

        Args:
            text_content: The text extracted from the document
            account_number: Optional account number
            file_type: Type of file (csv, pdf, excel)
            file_content: Optional binary file content (needed for Excel CSV retry)

        Returns:
            Dict with transactions and metadata
        """
        lines = text_content.split('\n')
        total_lines = len(lines)

        # PDFs shouldn't reach here - they use _parse_pdf_with_vision directly
        # But keep this as fallback in case text extraction is used
        if file_type == "pdf":
            print(f"📄 Processing PDF with {total_lines} lines using AI extraction (text fallback)...")
            if total_lines > 120:
                result = self._parse_large_file_in_chunks(text_content, account_number)
                if 'metadata' not in result:
                    result['metadata'] = {}
                result['metadata']['parsing_method'] = "pdf_text_chunked"
                return result
            else:
                result = self._parse_single_chunk(text_content, account_number)
                if 'metadata' not in result:
                    result['metadata'] = {}
                result['metadata']['parsing_method'] = "pdf_text_single"
                return result

        # CSV/Excel files use two-step approach
        print(f"📄 Processing {file_type.upper()} file with {total_lines} lines using two-step approach...")

        # Store line count in metadata for analytics
        file_line_count = total_lines

        # Step 1: Extract format specification from first 50 lines
        print("   Step 1: Analyzing file format with AI...")
        first_lines = '\n'.join(lines[:50])

        try:
            format_spec = self._extract_format_specification(first_lines)
            print(f"   ✓ Format detected: {format_spec['format']['delimiter']}-delimited, "
                  f"transactions start at row {format_spec['format']['transaction_start_row']}")
        except Exception as e:
            print(f"   ⚠️  Format extraction failed: {e}")
            print("   Falling back to old chunking method...")
            # Fallback to old method if format extraction fails
            if total_lines > 120:
                result = self._parse_large_file_in_chunks(text_content, account_number)
                if 'metadata' not in result:
                    result['metadata'] = {}
                result['metadata']['parsing_method'] = "format_detect_failed_chunked"
                return result
            else:
                result = self._parse_single_chunk(text_content, account_number)
                if 'metadata' not in result:
                    result['metadata'] = {}
                result['metadata']['parsing_method'] = "format_detect_failed_single"
                return result

        # Step 2: Parse full file with Python using format specification
        print("   Step 2: Parsing all transactions with Python...")
        result = self._parse_with_format_spec(text_content, format_spec)

        print(f"   ✓ Parsed {len(result['transactions'])} transactions")

        # Add token usage and format specification from format detection step
        if '_token_usage' in format_spec:
            if 'metadata' not in result:
                result['metadata'] = {}
            result['metadata']['input_tokens'] = format_spec['_token_usage']['input_tokens']
            result['metadata']['output_tokens'] = format_spec['_token_usage']['output_tokens']
            print(f"   ✓ Tokens: {format_spec['_token_usage']['input_tokens']:,} input, {format_spec['_token_usage']['output_tokens']:,} output")

        # Store format specification for debugging
        # Note: Don't store the metadata key from format_spec to avoid circular reference
        # since we're storing the format_spec inside result['metadata']
        if 'metadata' not in result:
            result['metadata'] = {}

        safe_format_spec = {}
        for key, value in format_spec.items():
            # Skip _token_usage (already stored separately) and metadata (would cause circular ref)
            if key in ('_token_usage', 'metadata'):
                continue
            safe_format_spec[key] = value

        result['metadata']['format_specification'] = safe_format_spec

        # If we got very few or zero transactions, the format spec might be wrong
        # This can happen with PDFs where text extraction creates multi-line entries
        # instead of tabular data. Fall back to AI-based extraction.
        if len(result['transactions']) == 0:
            print("   ⚠️  No transactions found - format spec may be incorrect")

            # For Excel files, try converting to CSV first before falling back to AI
            if file_type == "excel" and file_content is not None:
                print("   Trying CSV conversion as intermediate step...")
                try:
                    # Reset file pointer to beginning
                    file_content.seek(0)
                    csv_content = self._convert_excel_to_csv(file_content)
                    csv_lines = csv_content.split('\n')

                    print(f"   Step 1: Analyzing CSV format with AI...")
                    first_csv_lines = '\n'.join(csv_lines[:50])
                    csv_format_spec = self._extract_format_specification(first_csv_lines)
                    print(f"   ✓ CSV Format detected: {csv_format_spec['format']['delimiter']}-delimited")

                    print("   Step 2: Parsing CSV with Python...")
                    csv_result = self._parse_with_format_spec(csv_content, csv_format_spec)
                    print(f"   ✓ Parsed {len(csv_result['transactions'])} transactions from CSV")

                    if len(csv_result['transactions']) > 0:
                        # CSV conversion succeeded!
                        print("   ✅ CSV conversion successful - using CSV result")

                        # Add metadata about the conversion process
                        if 'metadata' not in csv_result:
                            csv_result['metadata'] = {}

                        # Store both format specs for debugging
                        csv_result['metadata']['format_specification'] = safe_format_spec  # Original tab-delimited
                        csv_result['metadata']['csv_format_specification'] = {
                            k: v for k, v in csv_format_spec.items()
                            if k not in ('_token_usage', 'metadata')
                        }
                        csv_result['metadata']['parsing_method'] = "format_spec_csv_retry"
                        csv_result['metadata']['file_line_count'] = file_line_count

                        # Accumulate token usage from both format detection attempts
                        if '_token_usage' in format_spec:
                            format_input = format_spec['_token_usage'].get('input_tokens', 0)
                            format_output = format_spec['_token_usage'].get('output_tokens', 0)
                            csv_result['metadata']['input_tokens'] = csv_result['metadata'].get('input_tokens', 0) + format_input
                            csv_result['metadata']['output_tokens'] = csv_result['metadata'].get('output_tokens', 0) + format_output

                        return csv_result
                    else:
                        print("   ⚠️  CSV conversion also found 0 transactions")

                except Exception as e:
                    print(f"   ⚠️  CSV conversion failed: {e}")

            print("   Falling back to AI-based extraction...")

            # Preserve the failed format spec for debugging
            failed_format_spec = safe_format_spec.copy()
            failed_format_spec['_fallback_reason'] = 'No transactions found with format spec'

            # Get result from AI extraction
            if total_lines > 120:
                fallback_result = self._parse_large_file_in_chunks(text_content, account_number)
                parsing_method = "format_spec_fallback_chunked"
            else:
                fallback_result = self._parse_single_chunk(text_content, account_number)
                parsing_method = "format_spec_fallback_single"

            # Add the failed format spec to metadata for debugging
            if 'metadata' not in fallback_result:
                fallback_result['metadata'] = {}
            fallback_result['metadata']['failed_format_specification'] = failed_format_spec
            fallback_result['metadata']['parsing_method'] = parsing_method
            fallback_result['metadata']['file_line_count'] = file_line_count

            # Preserve token usage from format detection step
            if '_token_usage' in format_spec:
                # Add format detection tokens to the AI extraction tokens
                format_input = format_spec['_token_usage'].get('input_tokens', 0)
                format_output = format_spec['_token_usage'].get('output_tokens', 0)
                fallback_result['metadata']['input_tokens'] = fallback_result['metadata'].get('input_tokens', 0) + format_input
                fallback_result['metadata']['output_tokens'] = fallback_result['metadata'].get('output_tokens', 0) + format_output

            return fallback_result

        # Format spec succeeded
        result['metadata']['parsing_method'] = "format_spec_python"
        result['metadata']['file_line_count'] = file_line_count
        return result

    def _parse_large_file_in_chunks(
        self,
        text_content: str,
        account_number: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Parse a large file by splitting into chunks and merging results

        Args:
            text_content: The full text content
            account_number: Optional account number

        Returns:
            Merged Dict with all transactions and metadata
        """
        chunks = self._chunk_text_by_lines(text_content, lines_per_chunk=100)
        print(f"   Split into {len(chunks)} chunks")

        all_transactions = []
        merged_metadata = {}
        total_input_tokens = 0
        total_output_tokens = 0

        for i, chunk in enumerate(chunks, 1):
            print(f"   Processing chunk {i}/{len(chunks)}...")
            try:
                chunk_result = self._parse_single_chunk(chunk, account_number)

                # Collect transactions
                all_transactions.extend(chunk_result.get("transactions", []))

                # Merge metadata (take from first chunk, update dates from all chunks)
                if not merged_metadata:
                    merged_metadata = chunk_result.get("metadata", {})
                else:
                    # Update date ranges
                    chunk_meta = chunk_result.get("metadata", {})
                    if chunk_meta.get("statement_start_date"):
                        if not merged_metadata.get("statement_start_date") or \
                           chunk_meta["statement_start_date"] < merged_metadata["statement_start_date"]:
                            merged_metadata["statement_start_date"] = chunk_meta["statement_start_date"]
                    if chunk_meta.get("statement_end_date"):
                        if not merged_metadata.get("statement_end_date") or \
                           chunk_meta["statement_end_date"] > merged_metadata["statement_end_date"]:
                            merged_metadata["statement_end_date"] = chunk_meta["statement_end_date"]
                    # Update closing balance from last chunk
                    if chunk_meta.get("closing_balance") is not None:
                        merged_metadata["closing_balance"] = chunk_meta["closing_balance"]

                # Accumulate token usage
                total_input_tokens += chunk_result.get("metadata", {}).get("input_tokens", 0)
                total_output_tokens += chunk_result.get("metadata", {}).get("output_tokens", 0)

            except Exception as e:
                print(f"   ⚠️  Chunk {i} failed: {e}")
                # Continue with other chunks

        # Update token usage in metadata
        merged_metadata["input_tokens"] = total_input_tokens
        merged_metadata["output_tokens"] = total_output_tokens

        print(f"✓ Processed all chunks. Total transactions: {len(all_transactions)}")

        return {
            "transactions": all_transactions,
            "metadata": merged_metadata
        }

    def _extract_format_specification(self, first_lines: str) -> Dict[str, Any]:
        """
        Analyze the first lines of a file to extract format specification.
        This uses AI to understand the file structure, then Python does the actual parsing.

        Args:
            first_lines: First ~50 lines of the file

        Returns:
            Dict containing format specification with:
            - metadata: account, currency, dates, balances
            - format: delimiter, quote_char, encoding, date_format
            - columns: mapping of column indices to field names
            - transaction_start_row: row number where transactions begin
            - amount_rules: how to interpret amounts (negative=debit, decimal separator, etc.)
        """
        prompt = """You are a financial document format analyzer. Analyze this sample from a bank statement and extract the file format specification.

Your task is to understand the structure and return a JSON specification that Python code can use to parse the full file.

Return ONLY valid JSON with this exact structure:

{
  "metadata": {
    "account_number": "account number (remove all spaces from IBANs)",
    "currency": "3-letter ISO code (USD/EUR/GBP/PLN) - look for symbols €,$,£,zł",
    "statement_start_date": "YYYY-MM-DD or null",
    "statement_end_date": "YYYY-MM-DD or null",
    "opening_balance": 1000.00 or null,
    "closing_balance": 500.00 or null
  },
  "format": {
    "delimiter": "," or ";" or "\\t" or "|",
    "quote_char": "\\"" or "'" or null,
    "has_header_row": true or false,
    "transaction_start_row": 15,
    "date_format": "%Y-%m-%d" or "%d.%m.%Y" or "%m/%d/%Y" etc,
    "decimal_separator": "." or ",",
    "thousands_separator": "," or "." or " " or null
  },
  "columns": {
    "date": 0,
    "amount": 8,
    "amount_fallback_columns": [10, 12],
    "description": 2,
    "description_parts": {
      "counterparty": 2,
      "title": 3
    },
    "reference": 7,
    "balance": null,
    "transaction_type": null,
    "currency": 9
  },
  "amount_rules": {
    "negative_means_debit": true,
    "separate_debit_credit_columns": false,
    "debit_column": null,
    "credit_column": null,
    "multiple_amount_columns": false,
    "check_all_amount_columns": [8, 10, 12]
  },
  "notes": "Any important observations about the format"
}

Field details:
- transaction_start_row: The row index (0-based) where the FIRST actual transaction data row begins.
  CRITICAL: This must point to the FIRST row containing actual transaction data, NOT a header/label row.
  Headers with "#" prefix or column names like "#Data operacji" are NOT transaction rows.
  Transaction rows typically start with dates (e.g., "2025-07-18;", "2024-10-06;").
  Count row numbers carefully from 0. If transactions start on line 28, set this to 27 (0-based).
- columns: Map field names to column indices (0-based). Use null if column doesn't exist
- columns.description: Primary description column index (will be used if description_parts is not specified)
- columns.description_parts: OPTIONAL object with counterparty and title column indices. Use this when the statement has separate columns for:
  - counterparty: The other party in the transaction (e.g., "Dane kontrahenta", "Payee", "Merchant")
  - title: The transaction title/purpose (e.g., "Tytuł", "Description", "Purpose")
  If both exist, they will be concatenated as: "counterparty title"
- columns.amount_fallback_columns: Array of alternative column indices to check if primary amount column is empty
- amount_rules.negative_means_debit: true if negative amounts = money out, false if positive = money out
- amount_rules.separate_debit_credit_columns: true if debits and credits are in different columns
- amount_rules.check_all_amount_columns: If multiple amount columns exist (like col 8, 10, 12), list ALL of them. Parser will use first non-empty value.

CRITICAL - Multiple Amount Columns:
Some bank statements have MULTIPLE amount columns (e.g., "Transaction Amount", "Block Amount", "Foreign Currency Amount").
Look at the actual DATA rows to see which column contains values for most transactions.
If different transactions use different amount columns, set amount_rules.check_all_amount_columns to list ALL possible amount columns.

IMPORTANT:
- Return ONLY the JSON, no other text
- Be precise with row/column indices (0-based)
- Look at ACTUAL DATA rows, not just headers - headers can be misleading
- If unsure which column has the amount, list multiple columns in check_all_amount_columns
- CRITICAL FOR DESCRIPTIONS: Check if the statement has SEPARATE columns for:
  1. The counterparty/payee/merchant (who you paid or who paid you)
  2. The transaction purpose/title/description (what the payment was for)
  If BOTH exist as separate columns, use description_parts to specify both column indices.
  If only one combined description column exists, use the simple description field.
  When in doubt, prefer description_parts if you see any separation of counterparty vs purpose.

CRITICAL - Finding transaction_start_row (MOST IMPORTANT - READ CAREFULLY):

Step-by-step algorithm to find the correct transaction_start_row:

STEP 1: Scan the file from top to bottom, looking at the FIRST column of each row
STEP 2: Find the FIRST row where the first column contains a DATE pattern (YYYY-MM-DD or similar)
STEP 3: Verify the row BEFORE that row does NOT contain a date (it should be a header like "#Data operacji" or blank)
STEP 4: That row number (0-based) is your transaction_start_row

COMMON MISTAKE TO AVOID:
- DO NOT skip the first date row just because the next row has the same date!
- DO NOT look for "date changes" - the FIRST date row is what counts!
- Multiple consecutive rows can have the SAME date - this is normal for same-day transactions

CONCRETE EXAMPLE (this is the EXACT pattern you'll see in mBank files):
Row 24: "PLN;301 132,84;-321 430,40;"          <- NO date in first column
Row 25: "#Data operacji;#Opis operacji;..."    <- NO date, this is a header (has # prefix)
Row 26: "2025-07-18;KAMIL GRYMUZA..."          <- ✓ FIRST DATE! Answer = 26
Row 27: "2025-07-18;KAMIL GRYMUZA..."          <- Same date, but DON'T start here
Row 28: "2025-06-13;..."                        <- Different date, but DON'T start here

Correct answer: transaction_start_row = 26

WRONG ANSWERS AND WHY:
- transaction_start_row = 27: WRONG! You skipped row 26 which has a valid transaction
- transaction_start_row = 28: WRONG! You skipped rows 26 and 27 which have valid transactions

VERIFICATION CHECKLIST (use this to verify your answer before responding):
✓ Row [transaction_start_row - 1] does NOT start with a date pattern (should be header or blank)
✓ Row [transaction_start_row] DOES start with a date pattern (YYYY-MM-DD)
✓ There is NO earlier row with a date in the first column
✓ You are including ALL transactions from this point forward (not skipping any date rows)

Sample file content:

"""

        # Call AI model to extract format
        if self.client:  # Claude
            model_name = "claude-sonnet-4-5-20250929" if self.model == "claude-sonnet" else "claude-3-5-haiku-20241022"

            # Optionally add timestamp to break caching (for testing)
            content = prompt + first_lines
            if self.disable_cache:
                import time
                cache_breaker = f"\n\n<!-- Analysis timestamp: {time.time()} -->\n"
                content = prompt + cache_breaker + first_lines

            response = self.client.messages.create(
                model=model_name,
                max_tokens=2048,
                temperature=0.0,  # Use deterministic mode for consistent results
                messages=[{"role": "user", "content": content}]
            )
            response_text = response.content[0].text
            input_tokens = response.usage.input_tokens
            output_tokens = response.usage.output_tokens
        else:  # OpenAI
            response = self.openai_client.chat.completions.create(
                model="gpt-4o",
                max_tokens=2048,
                messages=[{"role": "user", "content": prompt + first_lines}]
            )
            response_text = response.choices[0].message.content
            input_tokens = response.usage.prompt_tokens
            output_tokens = response.usage.completion_tokens

        # Parse JSON response
        try:
            # Extract JSON from response (might have markdown code blocks)
            import re
            json_match = re.search(r'```json\s*(\{.*?\})\s*```', response_text, re.DOTALL)
            if json_match:
                response_text = json_match.group(1)
            elif '```' in response_text:
                # Remove any code block markers
                response_text = re.sub(r'```[a-z]*\s*', '', response_text)
                response_text = response_text.replace('```', '')

            format_spec = json.loads(response_text.strip())

            # Add token usage to format spec so it can be tracked
            format_spec['_token_usage'] = {
                'input_tokens': input_tokens,
                'output_tokens': output_tokens
            }

            return format_spec
        except json.JSONDecodeError as e:
            print(f"Failed to parse format specification: {e}")
            print(f"Response: {response_text[:500]}")
            raise ValueError(f"AI returned invalid JSON: {str(e)}")

    def _parse_with_format_spec(
        self,
        text_content: str,
        format_spec: Dict[str, Any]
    ) -> Dict[str, Any]:
        """
        Parse the full file using Python based on the format specification.
        This is fast and has no token limits.

        Args:
            text_content: Full file content
            format_spec: Format specification from _extract_format_specification

        Returns:
            Dict with transactions and metadata
        """
        import re
        from datetime import datetime
        from io import StringIO

        lines = text_content.split('\n')
        transactions = []

        # Extract format details
        delimiter = format_spec['format']['delimiter']
        quote_char = format_spec['format'].get('quote_char')
        start_row = format_spec['format']['transaction_start_row']
        date_format = format_spec['format']['date_format']
        decimal_sep = format_spec['format'].get('decimal_separator', '.')
        thousands_sep = format_spec['format'].get('thousands_separator')

        # Column mappings
        col_map = format_spec['columns']
        amount_rules = format_spec['amount_rules']

        # Process transaction rows
        for i, line in enumerate(lines[start_row:], start=start_row):
            if not line.strip():
                continue

            try:
                # Parse CSV line
                if quote_char:
                    reader = csv.reader(StringIO(line), delimiter=delimiter, quotechar=quote_char)
                else:
                    reader = csv.reader(StringIO(line), delimiter=delimiter)

                row = next(reader)

                # Skip if not enough columns - flatten any list values for comparison
                max_col_indices = []
                for v in col_map.values():
                    if v is None:
                        continue
                    elif isinstance(v, dict):
                        # Handle nested dicts like description_parts
                        for nested_v in v.values():
                            if nested_v is not None and isinstance(nested_v, int):
                                max_col_indices.append(nested_v)
                    elif isinstance(v, list):
                        max_col_indices.extend([x for x in v if isinstance(x, int)])
                    elif isinstance(v, int):
                        max_col_indices.append(v)

                if max_col_indices and len(row) <= max(max_col_indices):
                    continue

                # Extract fields
                date_str = row[col_map['date']].strip() if col_map['date'] is not None else None

                # Extract description - support both simple and multi-part descriptions
                desc = ""
                if 'description_parts' in col_map and col_map['description_parts']:
                    # Multi-part description (e.g., counterparty + title)
                    parts = col_map['description_parts']
                    desc_components = []

                    if 'counterparty' in parts and parts['counterparty'] is not None and parts['counterparty'] < len(row):
                        counterparty = row[parts['counterparty']].strip()
                        if counterparty:
                            desc_components.append(counterparty)

                    if 'title' in parts and parts['title'] is not None and parts['title'] < len(row):
                        title = row[parts['title']].strip()
                        if title:
                            desc_components.append(title)

                    desc = " ".join(desc_components)
                elif col_map.get('description') is not None:
                    # Simple single-column description
                    desc = row[col_map['description']].strip() if col_map['description'] < len(row) else ""

                ref = row[col_map['reference']].strip() if col_map['reference'] is not None and col_map['reference'] < len(row) else None
                balance_str = row[col_map['balance']].strip() if col_map['balance'] is not None and col_map['balance'] < len(row) else None

                # Parse amount - handle multiple possible amount columns
                amount = None

                if amount_rules.get('separate_debit_credit_columns'):
                    debit_str = row[amount_rules['debit_column']].strip() if amount_rules['debit_column'] < len(row) else ""
                    credit_str = row[amount_rules['credit_column']].strip() if amount_rules['credit_column'] < len(row) else ""

                    if debit_str:
                        debit_str = debit_str.replace(thousands_sep, '') if thousands_sep else debit_str
                        debit_str = debit_str.replace(decimal_sep, '.')
                        amount = -abs(float(re.sub(r'[^\d.-]', '', debit_str)))
                    elif credit_str:
                        credit_str = credit_str.replace(thousands_sep, '') if thousands_sep else credit_str
                        credit_str = credit_str.replace(decimal_sep, '.')
                        amount = abs(float(re.sub(r'[^\d.-]', '', credit_str)))
                else:
                    # Check if we need to look in multiple amount columns
                    amount_columns = amount_rules.get('check_all_amount_columns', [col_map.get('amount')])
                    if not isinstance(amount_columns, list):
                        amount_columns = [amount_columns]

                    # Try each amount column until we find a non-empty value
                    for amt_col in amount_columns:
                        if amt_col is None or amt_col >= len(row):
                            continue

                        amount_str = row[amt_col].strip()
                        if amount_str and amount_str not in ['', '-', '0', '0,00', '0.00']:
                            try:
                                amount_str = amount_str.replace(thousands_sep, '') if thousands_sep else amount_str
                                amount_str = amount_str.replace(decimal_sep, '.')
                                amount = float(re.sub(r'[^\d.-]', '', amount_str))
                                break  # Found valid amount, stop looking
                            except (ValueError, AttributeError):
                                continue

                # Skip transaction if no valid amount found
                if amount is None:
                    continue

                # Parse date
                if date_str:
                    try:
                        date_obj = datetime.strptime(date_str, date_format)
                        date = date_obj.strftime('%Y-%m-%d')
                    except ValueError:
                        # Try common fallback formats
                        for fmt in ['%Y-%m-%d', '%d.%m.%Y', '%m/%d/%Y', '%d/%m/%Y']:
                            try:
                                date_obj = datetime.strptime(date_str, fmt)
                                date = date_obj.strftime('%Y-%m-%d')
                                break
                            except ValueError:
                                continue
                        else:
                            print(f"   Skipping transaction with invalid date: {date_str}")
                            continue
                else:
                    continue

                # Parse balance
                balance = None
                if balance_str:
                    try:
                        balance_str = balance_str.replace(thousands_sep, '') if thousands_sep else balance_str
                        balance_str = balance_str.replace(decimal_sep, '.')
                        balance = float(re.sub(r'[^\d.-]', '', balance_str))
                    except (ValueError, AttributeError):
                        pass

                # Determine transaction type
                if amount < 0:
                    txn_type = "DEBIT"
                elif amount > 0:
                    txn_type = "CREDIT"
                else:
                    txn_type = "OTHER"

                # Create transaction
                transactions.append({
                    "date": date,
                    "amount": amount,
                    "description": desc,
                    "transaction_type": txn_type,
                    "reference": ref if ref else None,
                    "balance": balance
                })

            except (IndexError, ValueError) as e:
                # Skip malformed rows silently
                continue

        return {
            "transactions": transactions,
            "metadata": format_spec['metadata']
        }

    def _parse_single_chunk(
        self,
        text_content: str,
        account_number: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Parse a single chunk of text content with Claude AI

        Args:
            text_content: The text to parse
            account_number: Optional account number

        Returns:
            Dict with transactions and metadata
        """
        prompt = """You are a financial document parser. Analyze the following document and extract ALL transaction data.

The document may be a bank statement, credit card statement, or any financial transaction record.

Your task:
1. Extract metadata first (one line starting with "METADATA,")
2. Then extract all transactions (one transaction per line starting with "TXN,")

Output format - CSV with properly quoted fields:

METADATA,account_number,currency,statement_start_date,statement_end_date,opening_balance,closing_balance
TXN,date,amount,"description",transaction_type,reference,balance
TXN,date,amount,"description",transaction_type,reference,balance
...

Example output:
METADATA,123456789,USD,2024-01-01,2024-01-31,2000.00,1949.75
TXN,2024-01-15,-50.25,"Amazon Purchase",DEBIT,REF123456,1949.75
TXN,2024-01-16,500.00,"Salary | Deposit",CREDIT,SAL2024,2449.75

Field details:
- date: ISO format YYYY-MM-DD
- amount: positive for credits/deposits, negative for debits/withdrawals
- description: transaction description/payee (MUST be quoted if contains commas or special chars)
- transaction_type: CREDIT, DEBIT, TRANSFER, FEE, INTEREST, or OTHER
- reference: transaction ID/reference number (empty if not available)
- balance: account balance after transaction (empty if not available)
- account_number: remove ALL whitespace/spaces from IBANs (e.g., "PL 27 1050..." → "PL271050...")
- currency: 3-letter ISO code (USD, EUR, GBP, PLN, etc.) - look for symbols (€,$,£,zł) or codes in document
- If currency not found, use EUR as default
- opening_balance/closing_balance: empty if not available

IMPORTANT:
- Return ONLY the CSV format, no other text or explanations
- ALWAYS quote description fields (they may contain commas, pipes, or other special characters)
- One METADATA line, then TXN lines for ALL transactions
- Be thorough - extract ALL transactions you find
- Amounts: negative for money going out, positive for money coming in

Document content:

"""

        # Call appropriate AI model
        if self.client:  # Claude
            model_name = "claude-sonnet-4-5-20250929" if self.model == "claude-sonnet" else "claude-3-5-haiku-20241022"
            response = self.client.messages.create(
                model=model_name,
                max_tokens=8192,
                messages=[
                    {
                        "role": "user",
                        "content": prompt + text_content
                    }
                ]
            )
            response_text = response.content[0].text
            input_tokens = response.usage.input_tokens
            output_tokens = response.usage.output_tokens

        else:  # OpenAI
            response = self.openai_client.chat.completions.create(
                model="gpt-4o",
                max_tokens=8192,
                messages=[
                    {
                        "role": "user",
                        "content": prompt + text_content
                    }
                ]
            )
            response_text = response.choices[0].message.content
            input_tokens = response.usage.prompt_tokens
            output_tokens = response.usage.completion_tokens

        # Parse pipe-delimited format
        try:
            result = self._parse_pipe_delimited_response(response_text)
        except Exception as e:
            # Save the raw response for debugging
            import tempfile
            debug_file = tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='_claude_response.txt')
            debug_file.write(response_text)
            debug_file.close()
            print(f"⚠️  Parse error. Raw response saved to: {debug_file.name}")
            raise ValueError(f"Could not parse Claude response: {str(e)}")

        # Override account number if provided
        if account_number:
            result["metadata"]["account_number"] = account_number

        # Add token usage to metadata for tracking
        if "metadata" not in result:
            result["metadata"] = {}
        result["metadata"]["input_tokens"] = input_tokens
        result["metadata"]["output_tokens"] = output_tokens

        return result

    def _parse_pipe_delimited_response(self, response_text: str) -> Dict[str, Any]:
        """
        Parse Claude's CSV response format into structured data

        Args:
            response_text: The raw response from Claude (CSV format)

        Returns:
            Dict with transactions and metadata
        """
        import csv
        from io import StringIO

        lines = response_text.strip().split('\n')
        transactions = []
        metadata = {}

        for line in lines:
            line = line.strip()
            if not line:
                continue

            # Use CSV reader to properly handle quoted fields
            try:
                reader = csv.reader(StringIO(line))
                parts = next(reader)
            except Exception as e:
                print(f"   Skipping malformed line: {line[:100]}")
                continue

            if not parts:
                continue

            # Check the first field to determine line type
            line_type = parts[0]

            if line_type == 'METADATA':
                # Parse: METADATA,account,currency,start_date,end_date,opening_bal,closing_bal
                if len(parts) >= 7:
                    metadata = {
                        "account_number": parts[1].strip() or None,
                        "currency": parts[2].strip() or "EUR",
                        "statement_start_date": parts[3].strip() or None,
                        "statement_end_date": parts[4].strip() or None,
                        "opening_balance": float(parts[5].strip()) if parts[5].strip() else None,
                        "closing_balance": float(parts[6].strip()) if parts[6].strip() else None
                    }

            elif line_type == 'TXN':
                # Parse: TXN,date,amount,description,type,reference,balance
                if len(parts) >= 5:  # At least TXN,date,amount,description,type
                    try:
                        txn = {
                            "date": parts[1].strip(),
                            "amount": float(parts[2].strip()),
                            "description": parts[3].strip(),
                            "transaction_type": parts[4].strip() if len(parts) > 4 else "OTHER",
                            "reference": parts[5].strip() if len(parts) > 5 and parts[5].strip() else None,
                            "balance": float(parts[6].strip()) if len(parts) > 6 and parts[6].strip() else None
                        }
                        transactions.append(txn)
                    except (ValueError, IndexError) as e:
                        # Skip malformed transaction lines
                        print(f"   Skipping malformed transaction: {line[:100]}")
                        continue

        # Empty statements (no transactions) are valid - return empty list
        if not transactions:
            print("   ℹ️  No transactions found (empty statement)")

        return {
            "transactions": transactions,
            "metadata": metadata
        }

    def convert_to_mt940(
        self,
        transactions_data: Dict[str, Any],
        account_number: Optional[str] = None
    ) -> str:
        """
        Convert parsed transactions to MT940 format

        Args:
            transactions_data: Dict with transactions and metadata from parse_document
            account_number: Override account number

        Returns:
            MT940 formatted string
        """
        metadata = transactions_data.get("metadata", {})
        transactions = transactions_data.get("transactions", [])

        # Use provided account number or from metadata
        acc_num = account_number or metadata.get("account_number") or "UNKNOWN"

        # Remove all whitespace from account number (IBANs should not have spaces)
        # BUT preserve any existing slashes
        acc_num = acc_num.replace(" ", "").replace("\t", "").replace("\n", "")

        # Add slash prefix for IBANs (SWIFT MT940 standard: :25:/IBAN)
        # IBAN format: 2 letters + 2 digits + up to 30 alphanumeric characters
        # Only add slash if account looks like IBAN and doesn't already have it
        acc_num_without_slash = acc_num.lstrip('/')
        if (acc_num_without_slash and
            len(acc_num_without_slash) >= 15 and
            acc_num_without_slash[:2].isalpha() and
            not acc_num.startswith('/')):
            # Add slash prefix for IBAN (best practice per SWIFT MT940 standard)
            acc_num = f"/{acc_num_without_slash}"

        # Get dates
        start_date = metadata.get("statement_start_date")
        end_date = metadata.get("statement_end_date")

        if not start_date and transactions:
            start_date = transactions[0]["date"]
        if not end_date and transactions:
            end_date = transactions[-1]["date"]

        # Format dates for MT940
        # Field 28C: Statement number (5 digits per SWIFT standard)
        statement_num = "00001"  # First statement (we're generating from a single document)
        opening_balance = metadata.get("opening_balance")
        closing_balance = metadata.get("closing_balance")

        # Calculate balances from transactions if not provided
        if opening_balance is None and transactions:
            # If we have first transaction balance, use it as opening
            first_txn_balance = transactions[0].get("balance")
            if first_txn_balance is not None:
                opening_balance = first_txn_balance - transactions[0]["amount"]
            else:
                opening_balance = 0.0
        elif opening_balance is None:
            opening_balance = 0.0

        if closing_balance is None and transactions:
            # Use last transaction balance if available
            last_txn_balance = transactions[-1].get("balance")
            if last_txn_balance is not None:
                closing_balance = last_txn_balance
            else:
                # Calculate from opening balance and all transactions
                closing_balance = opening_balance + sum(txn["amount"] for txn in transactions)
        elif closing_balance is None:
            closing_balance = 0.0

        # Build MT940
        mt940_lines = []

        # Generate Field 20: Transaction Reference Number (unique message identifier)
        # Format: DATE-FILENAME (max 16 chars per SWIFT standard)
        # This provides uniqueness by combining date with source file identifier
        # Example: 20241015-ABC123 or 241015-STATEMENT
        if start_date:
            # Use YYMMDD format to save space (vs YYYYMMDD)
            date_part = datetime.strptime(start_date, "%Y-%m-%d").strftime("%y%m%d")

            # Extract meaningful part from filename (if available)
            source_filename = metadata.get("source_filename", "")
            if source_filename:
                # Remove extension and take first 8 chars of basename
                basename = source_filename.rsplit('.', 1)[0]  # Remove extension
                # Keep only alphanumeric chars
                basename_clean = ''.join(c for c in basename if c.isalnum())
                filename_part = basename_clean[:8].upper() if basename_clean else "STMT"
            else:
                filename_part = "STMT"

            # Combine: DATE-FILE (e.g., 241015-STATEMENT, 241015-INGCORP)
            field_20_ref = f"{date_part}-{filename_part}"

            # Ensure max 16 characters (SWIFT requirement)
            field_20_ref = field_20_ref[:16]
        else:
            # Fallback if no date available - use timestamp
            field_20_ref = datetime.now().strftime("%y%m%d%H%M%S")[:16]

        # Validate that we have minimum required data for valid MT940
        if not start_date or not end_date:
            raise EmptyStatementError(
                "The uploaded file does not contain valid transaction data or statement dates. "
                "Please ensure the file contains a bank statement with transaction information."
            )

        # Header
        mt940_lines.append(f":20:{field_20_ref}")
        mt940_lines.append(f":25:{acc_num}")
        mt940_lines.append(f":28C:{statement_num}")

        # Opening balance
        if start_date:
            start_date_formatted = datetime.strptime(start_date, "%Y-%m-%d").strftime("%y%m%d")
            opening_sign = "C" if opening_balance >= 0 else "D"
            opening_amount = abs(opening_balance)
            # Use comma for decimal separator (European MT940 format)
            opening_amount_str = f"{opening_amount:.2f}".replace('.', ',')
            # Get currency from metadata, default to EUR
            currency = metadata.get("currency", "EUR")
            mt940_lines.append(f":60F:{opening_sign}{start_date_formatted}{currency}{opening_amount_str}")

        # Transactions
        for idx, txn in enumerate(transactions, 1):
            date_obj = datetime.strptime(txn["date"], "%Y-%m-%d")
            value_date = date_obj.strftime("%y%m%d")  # YYMMDD format
            entry_date = date_obj.strftime("%m%d")     # MMDD format (optional entry date)

            amount = txn["amount"]
            debit_credit = "D" if amount < 0 else "C"
            amount_abs = abs(amount)

            # Format amount with comma (European format for MT940)
            amount_str = f"{amount_abs:.2f}".replace('.', ',')

            # Transaction type code
            # Format: [funds_code][3-char type code]
            # N = non-urgent, MSC = miscellaneous, TRF = transfer
            type_code = "NMSC"  # N=funds code, MSC=miscellaneous

            # Statement line format: :61:YYMMDD[MMDD][D|C]amount,decimals[N]TYPE
            # Value date (6) + Entry date (4) + Debit/Credit + Amount + Type code
            mt940_lines.append(f":61:{value_date}{entry_date}{debit_credit}{amount_str}{type_code}")

            # Transaction details - Field 86: max 6 lines of 65 characters each
            description = txn.get("description", "").replace("\n", " ")
            # Sanitize whitespace: replace multiple spaces with single space
            description = re.sub(r'\s+', ' ', description).strip()

            # If description is empty, use "-" as placeholder
            if not description:
                description = "-"

            # Split long descriptions across multiple lines (SWIFT MT940 spec: 6*65x)
            max_lines = 6
            line_length = 65
            desc_lines = []

            for i in range(0, min(len(description), max_lines * line_length), line_length):
                desc_lines.append(description[i:i+line_length])

            # First line starts with :86:, continuation lines don't have the tag
            mt940_lines.append(f":86:{desc_lines[0]}")
            for continuation_line in desc_lines[1:]:
                mt940_lines.append(continuation_line)

        # Closing balance
        if end_date:
            end_date_formatted = datetime.strptime(end_date, "%Y-%m-%d").strftime("%y%m%d")
            closing_sign = "C" if closing_balance >= 0 else "D"
            closing_amount = abs(closing_balance)
            # Use comma for decimal separator (European MT940 format)
            closing_amount_str = f"{closing_amount:.2f}".replace('.', ',')
            # Get currency from metadata, default to EUR
            currency = metadata.get("currency", "EUR")
            mt940_lines.append(f":62F:{closing_sign}{end_date_formatted}{currency}{closing_amount_str}")

        return "\n".join(mt940_lines)
